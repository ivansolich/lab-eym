from lab3.analisisResistividad import sigmaRho1,sigmaRho2,sigmaRho3
from openpyxl import Workbook

x1 = sigmaRho1
x2 = sigmaRho2
x3 = sigmaRho3

wb = Workbook()
ruta = 'salida.xlsx'

hoja = wb.active
hoja.title = "Fecha-Valor"

fila = 1 #Fila donde empezamos
col_x = 1 #Columna donde guardamos las fechas
col_y = 2 #Columna donde guardamos el dato asociados a cada fecha
col_z = 3

for x, y,z in zip(x1, x2,x3):
    hoja.cell(column=col_x, row=fila, value=x)
    hoja.cell(column=col_y, row=fila, value=y)
    hoja.cell(column=col_z, row=fila, value=z)
    fila+=1

wb.save(filename = ruta)